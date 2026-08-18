"""Testes do recorte visualizável (``censo amostra``).

O que precisa ficar fixado aqui é o que o usuário não consegue verificar
sozinho numa máquina que não abre o arquivo: que a leitura para cedo (e por
isso não estoura a memória), que os filtros não escondem a coluna que os
motivou, e que o .xlsx sai com encoding e tipos certos.
"""

from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import load_workbook

from censo_escolar import amostra
from censo_escolar.config import ENCODING, SEPARADOR

CABECALHO = "CO_ENTIDADE;NO_ENTIDADE;SG_UF;CO_MUNICIPIO;QT_MAT_BAS"


def _csv(tmp_path, linhas: int = 50):
    caminho = tmp_path / "microdados_ed_basica_2023.csv"
    corpo = [
        f"{1000 + i};ESCOLA {'MUNICIPAL' if i % 2 else 'ESTADUAL'} Nº {i} - AÇÃO;"
        f"{'MG' if i % 3 else 'SP'};3100{i:03d};{i * 7}"
        for i in range(linhas)
    ]
    caminho.write_text("\n".join([CABECALHO, *corpo]) + "\n", encoding=ENCODING)
    return caminho


def test_recorta_o_comeco_sem_ler_o_arquivo_todo(tmp_path, monkeypatch):
    csv = _csv(tmp_path, linhas=500)
    monkeypatch.setattr(amostra, "TAMANHO_BLOCO", 10)

    recorte, origem, lidas = amostra.amostrar(arquivo=csv, linhas=5)

    assert len(recorte) == 5
    assert origem == csv
    # O ponto do módulo: parou no primeiro bloco, não varreu as 500 linhas.
    assert lidas == 10


def test_filtro_exato_ignora_caixa(tmp_path):
    recorte, _, _ = amostra.amostrar(arquivo=_csv(tmp_path), linhas=10, onde={"SG_UF": "sp"})
    assert set(recorte["SG_UF"]) == {"SP"}


def test_contem_busca_no_meio_do_texto(tmp_path):
    recorte, _, _ = amostra.amostrar(
        arquivo=_csv(tmp_path), linhas=3, contem={"NO_ENTIDADE": "municipal"}
    )
    assert len(recorte) == 3
    assert all("MUNICIPAL" in nome for nome in recorte["NO_ENTIDADE"])


def test_coluna_do_filtro_entra_mesmo_sem_ter_sido_pedida(tmp_path):
    recorte, _, _ = amostra.amostrar(
        arquivo=_csv(tmp_path), linhas=3, colunas=["NO_ENTIDADE"], onde={"SG_UF": "SP"}
    )
    assert list(recorte.columns) == ["NO_ENTIDADE", "SG_UF"]


def test_sem_linhas_diz_quanto_varreu(tmp_path):
    with pytest.raises(amostra.SemLinhas, match="50 linhas varridas"):
        amostra.amostrar(arquivo=_csv(tmp_path), linhas=5, onde={"SG_UF": "ZZ"})


def test_coluna_de_filtro_inexistente_e_erro_com_dica(tmp_path):
    with pytest.raises(KeyError, match="censo colunas"):
        amostra.amostrar(arquivo=_csv(tmp_path), onde={"NAO_EXISTE": "1"})


def test_xlsx_sai_pronto_para_leitura(tmp_path):
    recorte, _, _ = amostra.amostrar(arquivo=_csv(tmp_path), linhas=4)
    destino = amostra.salvar(recorte, tmp_path / "amostra.xlsx")

    planilha = load_workbook(destino).active
    assert planilha.freeze_panes == "A2"  # cabeçalho à vista ao rolar
    assert planilha.auto_filter.ref == planilha.dimensions  # a busca visual
    assert planilha["A1"].value == "CO_ENTIDADE"
    assert planilha["A1"].font.bold
    # latin-1 do INEP lido e regravado sem estragar o acento.
    assert "AÇÃO" in planilha["B2"].value
    # Colunas numéricas viram número, para somar e ordenar na planilha.
    assert planilha["E2"].value == 0
    assert isinstance(planilha["A2"].value, int)


def test_zero_a_esquerda_continua_texto(tmp_path):
    """Código com zero à esquerda é identificador; virar número o destruiria."""
    csv = tmp_path / "codigos.csv"
    csv.write_text("CO_ORGAO_REGIONAL;QT\n0MI11;3\n00042;4\n", encoding=ENCODING)

    recorte, _, _ = amostra.amostrar(arquivo=csv, linhas=10)
    destino = amostra.salvar(recorte, tmp_path / "codigos.xlsx")

    planilha = load_workbook(destino).active
    assert planilha["A3"].value == "00042"
    assert planilha["B2"].value == 3


def test_csv_de_saida_sai_com_bom(tmp_path):
    recorte, _, _ = amostra.amostrar(arquivo=_csv(tmp_path), linhas=2)
    destino = amostra.salvar(recorte, tmp_path / "recorte.csv")

    assert destino.read_bytes().startswith(b"\xef\xbb\xbf")
    de_volta = pd.read_csv(destino, sep=SEPARADOR, encoding="utf-8-sig")
    assert len(de_volta) == 2


@pytest.mark.parametrize(
    ("entrada", "esperado"),
    [
        (["SG_UF=MG"], {"SG_UF": "MG"}),
        (["sg_uf = MG "], {"SG_UF": "MG"}),
        # Só o primeiro `=` separa: o valor pode conter outros.
        (["NO_ENTIDADE=A=B"], {"NO_ENTIDADE": "A=B"}),
    ],
)
def test_pares_de_filtro(entrada, esperado):
    assert amostra.pares_de_filtro(entrada, "onde") == esperado


def test_pares_de_filtro_sem_igual_e_erro():
    with pytest.raises(ValueError, match="COLUNA=VALOR"):
        amostra.pares_de_filtro(["SG_UF"], "onde")
