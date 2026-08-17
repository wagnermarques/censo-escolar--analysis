"""Esquema dos microdados: que colunas existem em cada ano, de que tipo, e o
dicionário que descreve as comuns.

Uma série histórica do Censo Escolar só é confiável sobre as variáveis que
existem em *todos* os anos comparados — e com o *mesmo tipo* em todos eles. As
duas condições falham calada:

- O INEP acrescenta, renomeia e aposenta variáveis a cada edição. Quem empilha
  anos com ``carregar_anos`` recebe ``NaN`` onde a coluna não existia, e um
  ``NaN`` somado vira zero na hora errada — o gráfico despenca e a causa é o
  arquivo, não a realidade.
- Um mesmo nome pode mudar de tipo entre anos. ``CO_ORGAO_REGIONAL`` é numérico
  em uns estados e alfanumérico (``0MI11``) em outros; concatenar
  ``Int64`` com ``string`` produz uma coluna ``object`` cheia de surpresa.

Este módulo responde às duas perguntas *antes* da análise: :func:`colunas_comuns`
resolve a primeira lendo só os cabeçalhos, e :func:`perfilar` resolve a segunda
lendo os dados de verdade — não o prefixo do nome, que mente. O produto final é
:func:`dicionario_de_dados`, uma tabela com uma linha por variável comum, já
com a descrição oficial do INEP e o dtype que serve a todos os anos.
"""

from __future__ import annotations

import re
from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from pathlib import Path

import pandas as pd

from censo_escolar.config import ENCODING, SEPARADOR, Paths, get_paths
from censo_escolar.loading import _dtypes_para, colunas_disponiveis, localizar_csv_escolas

#: O que cada prefixo do INEP significa. Serve à leitura humana do dicionário —
#: o tipo vem dos dados, não daqui (veja :func:`perfilar`).
NATUREZA_POR_PREFIXO: dict[str, str] = {
    "NU_": "número",
    "CO_": "código",
    "NO_": "nome",
    "SG_": "sigla",
    "DS_": "descrição",
    "TP_": "categórica codificada",
    "IN_": "indicador (0/1)",
    "QT_": "contagem",
    "DT_": "data",
    "TX_": "texto livre",
}

#: Ordem de promoção de dtypes: ao empilhar anos, o tipo que serve a todos é o
#: mais largo entre os observados. ``string`` é o topo porque aceita qualquer
#: coisa — inclusive o ``0MI11`` que fez um ``Int64`` falhar em outro ano.
ORDEM_DTYPE: tuple[str, ...] = ("Int8", "Int16", "Int32", "Int64", "Float64", "string")

#: Marcador para a coluna que existe no ano mas está inteiramente vazia. Não é
#: um dtype: é a ausência de evidência, e por isso não participa da promoção.
VAZIO = "vazio"

#: Quantos valores distintos guardar por coluna antes de desistir da contagem
#: exata. Cobre com folga toda ``TP_``/``IN_`` (o caso em que os distintos
#: interessam) sem carregar as ~200 mil razões sociais de ``NO_ENTIDADE``.
LIMITE_DISTINTOS = 200

#: Limites dos inteiros com sinal, do mais estreito ao mais largo.
_FAIXAS_INT: tuple[tuple[str, int, int], ...] = (
    ("Int8", -128, 127),
    ("Int16", -32_768, 32_767),
    ("Int32", -2**31, 2**31 - 1),
)

_ANO_NO_NOME = re.compile(r"(\d{4})$")

#: Nome de variável do INEP: maiúsculas, dígitos e sublinhado. Filtra as linhas
#: de nota de rodapé e de cabeçalho que sobram na planilha do dicionário.
_NOME_DE_VARIAVEL = re.compile(r"^[A-Z][A-Z0-9_]*$")


# --------------------------------------------------------------------------
# Que colunas existem em cada ano
# --------------------------------------------------------------------------


def anos_disponiveis(*, paths: Paths | None = None) -> list[int]:
    """Anos já extraídos em ``data/interim/`` cujo CSV de escolas é localizável.

    Um diretório extraído pela metade não entra na lista: o objetivo é que o
    resultado possa ser passado direto para as funções abaixo sem estourar.
    """
    paths = paths or get_paths()
    if not paths.interim.exists():
        return []

    anos: list[int] = []
    for diretorio in paths.interim.iterdir():
        casa = _ANO_NO_NOME.search(diretorio.name)
        if not (diretorio.is_dir() and casa):
            continue
        ano = int(casa.group(1))
        try:
            localizar_csv_escolas(ano, paths=paths)
        except FileNotFoundError:
            continue
        anos.append(ano)
    return sorted(anos)


def matriz_presenca(anos: Sequence[int], *, paths: Paths | None = None) -> pd.DataFrame:
    """Uma linha por variável, uma coluna por ano, ``True`` onde ela existe.

    Lê apenas os cabeçalhos dos CSVs, então roda em segundos mesmo com todos os
    anos. As linhas saem na ordem do arquivo mais recente (com as variáveis
    aposentadas logo depois), que agrupa as famílias ``QT_MAT_*``, ``IN_*`` etc.
    melhor do que a ordem alfabética.
    """
    anos = sorted(anos)
    if not anos:
        raise ValueError("informe ao menos um ano")

    por_ano = {ano: colunas_disponiveis(ano, paths=paths) for ano in anos}
    conjuntos = {ano: set(colunas) for ano, colunas in por_ano.items()}

    ordem: list[str] = []
    vistas: set[str] = set()
    for ano in reversed(anos):
        for coluna in por_ano[ano]:
            if coluna not in vistas:
                vistas.add(coluna)
                ordem.append(coluna)

    return pd.DataFrame(
        {ano: [coluna in conjuntos[ano] for coluna in ordem] for ano in anos},
        index=pd.Index(ordem, name="coluna"),
    )


def contagem_por_ano(matriz: pd.DataFrame) -> pd.DataFrame:
    """Quantas variáveis cada ano traz, a partir de :func:`matriz_presenca`."""
    return pd.DataFrame({"ano": list(matriz.columns), "variaveis": matriz.sum().to_numpy()})


def colunas_comuns(anos: Sequence[int], *, paths: Paths | None = None) -> list[str]:
    """As variáveis presentes em *todos* os anos informados."""
    matriz = matriz_presenca(anos, paths=paths)
    return list(matriz.index[matriz.all(axis=1)])


def resumo_presenca(anos: Sequence[int], *, paths: Paths | None = None) -> pd.DataFrame:
    """Em que anos cada variável aparece, e se aparece sem buracos.

    A coluna ``continua`` é a que separa duas situações bem diferentes: uma
    variável criada em 2021 (buraco só no começo, série histórica curta porém
    íntegra) e uma variável que sumiu num ano do meio (série que parece
    completa e não é).
    """
    matriz = matriz_presenca(anos, paths=paths)
    todos = list(matriz.columns)

    linhas = []
    for coluna, presenca in matriz.iterrows():
        presentes = [ano for ano in todos if bool(presenca[ano])]
        ausentes = [ano for ano in todos if not presenca[ano]]
        faixa = range(todos.index(presentes[0]), todos.index(presentes[-1]) + 1)
        linhas.append(
            {
                "coluna": coluna,
                "n_anos": len(presentes),
                "comum": len(ausentes) == 0,
                "primeiro_ano": presentes[0],
                "ultimo_ano": presentes[-1],
                "continua": len(presentes) == len(faixa),
                "anos_presentes": ", ".join(str(a) for a in presentes),
                "anos_ausentes": ", ".join(str(a) for a in ausentes),
            }
        )
    return pd.DataFrame(linhas)


# --------------------------------------------------------------------------
# De que tipo é cada variável, segundo os dados
# --------------------------------------------------------------------------


@dataclass
class _Acumulador:
    """Estatísticas de uma coluna, atualizadas bloco a bloco.

    Guarda resumos, nunca os valores: é o que permite varrer o arquivo inteiro
    (~215 mil linhas x ~400 colunas) sem que a memória cresça com o tamanho do
    CSV.
    """

    n: int = 0
    nulos: int = 0
    numerico: bool = True
    inteiro: bool = True
    zeros_a_esquerda: bool = False
    minimo: float | None = None
    maximo: float | None = None
    largura_max: int = 0
    distintos: set[str] = field(default_factory=set)
    truncado: bool = False


def _acumular(acc: _Acumulador, serie: pd.Series) -> None:
    acc.n += len(serie)
    valores = serie.dropna()
    acc.nulos += len(serie) - len(valores)
    if valores.empty:
        return

    unicos = valores.unique()
    acc.largura_max = max(acc.largura_max, max(len(v) for v in unicos))

    if not acc.truncado:
        acc.distintos.update(unicos.tolist())
        if len(acc.distintos) > LIMITE_DISTINTOS:
            acc.truncado = True
            acc.distintos = set(sorted(acc.distintos)[:LIMITE_DISTINTOS])

    # Uma vez que a coluna se revelou textual, não há o que reavaliar — e pular
    # esta parte é o que mantém o custo baixo justamente nas colunas caras
    # (``NO_ENTIDADE`` tem dezenas de milhares de valores distintos por bloco).
    if not acc.numerico:
        return

    numeros = pd.to_numeric(pd.Series(unicos), errors="coerce")
    if numeros.isna().any():
        acc.numerico = False
        return

    minimo, maximo = float(numeros.min()), float(numeros.max())
    acc.minimo = minimo if acc.minimo is None else min(acc.minimo, minimo)
    acc.maximo = maximo if acc.maximo is None else max(acc.maximo, maximo)
    if acc.inteiro and not bool((numeros % 1 == 0).all()):
        acc.inteiro = False
    # "00009" e 9 são o mesmo número e códigos diferentes. Quem tem zero à
    # esquerda é identificador, não quantidade, e só sobrevive como texto.
    if any(len(v) > 1 and v[0] == "0" and v[1] != "." for v in unicos):
        acc.zeros_a_esquerda = True


def _tipo_dos_dados(acc: _Acumulador) -> str:
    """O dtype mais econômico que representa o que foi observado, sem perda."""
    if acc.n == acc.nulos:
        return VAZIO
    if not acc.numerico or acc.zeros_a_esquerda:
        return "string"
    if not acc.inteiro:
        return "Float64"
    if acc.minimo is None or acc.maximo is None:
        return VAZIO
    for nome, minimo, maximo in _FAIXAS_INT:
        if acc.minimo >= minimo and acc.maximo <= maximo:
            return nome
    return "Int64"


def perfilar(
    ano: int,
    colunas: Sequence[str] | None = None,
    *,
    linhas: int | None = None,
    tamanho_bloco: int = 100_000,
    paths: Paths | None = None,
) -> pd.DataFrame:
    """Lê o CSV de ``ano`` como texto e descreve cada coluna a partir dos valores.

    Ler como texto é o ponto: qualquer dtype declarado na leitura já embutiria a
    resposta que estamos tentando obter, e um ``errors="coerce"`` esconderia
    exatamente os valores que interessam.

    :param colunas: quais perfilar. Por padrão, todas as do ano.
    :param linhas: lê só as ``linhas`` primeiras (ensaio rápido). Cuidado: o
        arquivo do INEP vem ordenado por UF, então um recorte do começo só
        enxerga os primeiros estados — foi assim que ``CO_ORGAO_REGIONAL``
        passou por numérico até alguém carregar um ano inteiro. Para decidir
        tipo, prefira a varredura completa.
    :param tamanho_bloco: linhas por bloco de leitura; controla memória, não
        resultado.

    Devolve uma linha por coluna, com ``dtype_dados`` (o que os valores pedem) ao
    lado de ``dtype_prefixo`` (o que a heurística de ``loading.py`` escolheria) e
    ``conflito`` marcando onde os dois discordam.
    """
    paths = paths or get_paths()
    csv = localizar_csv_escolas(ano, paths=paths)
    existentes = colunas_disponiveis(ano, paths=paths)
    alvo = [c for c in (colunas if colunas is not None else existentes) if c in set(existentes)]
    if not alvo:
        raise KeyError(f"nenhuma das colunas pedidas existe em {ano}")

    acumuladores: dict[str, _Acumulador] = {coluna: _Acumulador() for coluna in alvo}
    leitor = pd.read_csv(
        csv,
        sep=SEPARADOR,
        encoding=ENCODING,
        usecols=alvo,
        dtype=str,
        chunksize=tamanho_bloco,
        nrows=linhas,
    )
    for bloco in leitor:
        for coluna in alvo:
            _acumular(acumuladores[coluna], bloco[coluna])

    prefixos = _dtypes_para(alvo)
    linhas_saida = []
    for coluna in alvo:
        acc = acumuladores[coluna]
        tipo = _tipo_dos_dados(acc)
        dtype_prefixo = prefixos.get(coluna, "")
        linhas_saida.append(
            {
                "ano": ano,
                "coluna": coluna,
                "natureza": natureza(coluna),
                "dtype_dados": tipo,
                "dtype_prefixo": dtype_prefixo,
                # Conflito é só o que faz a heurística *perder* informação — o
                # ponto em que a leitura com dtypes declarados cai no modo
                # tolerante de `loading._ler_csv_tolerante`. Um prefixo largo
                # demais (``CO_UF`` como Int64 quando Int8 basta) desperdiça
                # memória, não dados, e aparece só na diferença entre as duas
                # colunas acima.
                "conflito": (
                    bool(dtype_prefixo)
                    and tipo != VAZIO
                    and _largura(tipo) > _largura(dtype_prefixo)
                ),
                "n_linhas": acc.n,
                "pct_nulo": (acc.nulos / acc.n) if acc.n else 1.0,
                "n_distintos": len(acc.distintos),
                "contagem_exata": not acc.truncado,
                "minimo": acc.minimo,
                "maximo": acc.maximo,
                "largura_max": acc.largura_max,
                "zeros_a_esquerda": acc.zeros_a_esquerda,
                "exemplos": ", ".join(sorted(acc.distintos)[:5]),
            }
        )
    return pd.DataFrame(linhas_saida)


def natureza(coluna: str) -> str:
    """O que o prefixo do nome diz que a variável é (só para leitura humana)."""
    return NATUREZA_POR_PREFIXO.get(coluna[:3], "")


def perfilar_anos(
    anos: Sequence[int],
    colunas: Sequence[str] | None = None,
    *,
    linhas: int | None = None,
    paths: Paths | None = None,
) -> pd.DataFrame:
    """Empilha :func:`perfilar` de vários anos, com a coluna ``ano`` distinguindo."""
    return pd.concat(
        [perfilar(ano, colunas, linhas=linhas, paths=paths) for ano in sorted(anos)],
        ignore_index=True,
    )


def tipos_por_ano(perfis: pd.DataFrame) -> pd.DataFrame:
    """Tabela cruzada coluna x ano com o ``dtype_dados`` observado em cada um."""
    return perfis.pivot(index="coluna", columns="ano", values="dtype_dados")


def promover(tipos: Iterable[str]) -> str:
    """O dtype mais estreito que comporta todos os ``tipos`` observados.

    ``VAZIO`` é ignorado: um ano em que a coluna veio toda vazia não é evidência
    de tipo nenhum, e deixá-lo participar rebaixaria a decisão sem motivo.
    """
    observados = [t for t in tipos if t != VAZIO and pd.notna(t)]
    if not observados:
        return VAZIO
    return max(observados, key=_largura)


def _largura(tipo: str) -> int:
    """Posição do dtype na ordem de promoção; desconhecido conta como o mais largo."""
    return ORDEM_DTYPE.index(tipo) if tipo in ORDEM_DTYPE else len(ORDEM_DTYPE)


# --------------------------------------------------------------------------
# A descrição oficial, da planilha que vem no próprio ZIP
# --------------------------------------------------------------------------


def localizar_dicionario_inep(ano: int, *, paths: Paths | None = None) -> Path:
    """Encontra a planilha ``ANEXO I - Dicionário de Dados`` do ano extraído."""
    paths = paths or get_paths()
    raiz = paths.interim / f"microdados_censo_escolar_{ano}"
    candidatos = [
        arquivo
        for arquivo in raiz.rglob("*.xlsx")
        # `~$...` são os arquivos de bloqueio que o Excel deixa para trás; eles
        # acompanham o ZIP de alguns anos e não são planilhas válidas.
        if not arquivo.name.startswith("~$") and "dicion" in arquivo.name.lower()
    ]
    if not candidatos:
        raise FileNotFoundError(
            f"Planilha de dicionário do INEP não encontrada sob {raiz}.\n"
            f"Rode: censo obter {ano}"
        )
    return max(candidatos, key=lambda p: p.stat().st_size)


def _aba_de_escolas(nomes: list[str]) -> str:
    """Escolhe a aba do arquivo de escolas entre as abas da planilha.

    O nome mudou a cada reforma do pacote — ``Cadastro_Escolas`` até 2022,
    ``microdados_unidade_coleta`` em 2023/2024, ``Tabela_de_Escola`` em 2025 —,
    e desde 2025 há uma aba por entidade. A de escolas é a primeira em todos os
    anos conferidos; os padrões abaixo evitam depender só disso.
    """
    for padrao in (r"escola", r"cadastro", r"unidade_coleta"):
        for nome in nomes:
            if re.search(padrao, nome, re.IGNORECASE) and not re.search(
                r"gestor", nome, re.IGNORECASE
            ):
                return nome
    return nomes[0]


def ler_dicionario_inep(
    ano: int,
    *,
    aba: str | None = None,
    paths: Paths | None = None,
) -> pd.DataFrame:
    """Lê a planilha de dicionário do INEP e devolve nome, descrição e domínio.

    A planilha não é uma tabela limpa: tem título, legenda de cores e um
    cabeçalho de duas alturas antes dos dados. Em vez de fixar o número de
    linhas a pular — que muda de ano para ano —, procuramos a linha que contém
    "Nome da Variável" e mapeamos as colunas pelo texto do cabeçalho.

    :param aba: força uma aba específica (útil para as tabelas de matrícula,
        docente e turma que surgiram em 2025).
    """
    import openpyxl

    caminho = localizar_dicionario_inep(ano, paths=paths)
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    try:
        ws = wb[aba] if aba else wb[_aba_de_escolas(wb.sheetnames)]
        linhas = [
            [str(c).strip() if c is not None else "" for c in linha]
            for linha in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()

    #: Campo do resultado -> como o rótulo dele começa na planilha. O casamento
    #: é por prefixo porque o INEP varia o resto: "Tam.(1)", "Tamanho",
    #: "Descrição da Variável" com e sem quebra de linha no meio.
    campos = (
        ("coluna", "nome da vari"),
        ("descricao", "descri"),
        ("tipo_inep", "tipo"),
        ("tamanho_inep", "tam"),
        ("categorias", "categoria"),
    )

    posicoes: dict[str, int] = {}
    inicio = 0
    for i, linha in enumerate(linhas):
        rotulos = [(texto.lower(), j) for j, texto in enumerate(linha) if texto]
        achados = {
            campo: pos
            for campo, prefixo in campos
            for rotulo, pos in rotulos
            if rotulo.startswith(prefixo)
        }
        if "coluna" not in achados:
            continue
        posicoes = achados
        inicio = i + 1
        break
    if not posicoes:
        raise ValueError(f"Cabeçalho 'Nome da Variável' não encontrado em {caminho}")

    def celula(linha: list[str], chave: str) -> str:
        pos = posicoes.get(chave)
        if pos is None or pos >= len(linha):
            return ""
        # As categorias vêm como lista de várias linhas dentro de uma célula;
        # achatamos para caber numa linha de CSV sem perder os rótulos.
        return " | ".join(p.strip() for p in linha[pos].split("\n") if p.strip())

    registros: dict[str, dict[str, str]] = {}
    for linha in linhas[inicio:]:
        if not linha or posicoes["coluna"] >= len(linha):
            continue
        nome = linha[posicoes["coluna"]]
        if not _NOME_DE_VARIAVEL.match(nome) or nome in registros:
            continue
        registros[nome] = {
            "coluna": nome,
            "descricao": celula(linha, "descricao"),
            "tipo_inep": celula(linha, "tipo_inep"),
            "tamanho_inep": celula(linha, "tamanho_inep"),
            "categorias": celula(linha, "categorias"),
        }

    return pd.DataFrame(list(registros.values()))


def descricoes(anos: Sequence[int], *, paths: Paths | None = None) -> pd.DataFrame:
    """Junta os dicionários do INEP de vários anos, o mais recente tendo a palavra.

    A redação das descrições é revisada entre edições; vale a do ano mais novo em
    que a variável aparece. Anos cuja planilha não abra são ignorados — perder a
    descrição é aceitável, travar o dicionário inteiro não.
    """
    partes = []
    for ano in sorted(anos, reverse=True):
        try:
            partes.append(ler_dicionario_inep(ano, paths=paths))
        except (FileNotFoundError, ValueError, ImportError):
            continue
    if not partes:
        return pd.DataFrame(
            columns=["coluna", "descricao", "tipo_inep", "tamanho_inep", "categorias"]
        )
    return pd.concat(partes, ignore_index=True).drop_duplicates("coluna", keep="first")


# --------------------------------------------------------------------------
# O dicionário de dados
# --------------------------------------------------------------------------

#: Arquivo padrão do dicionário gerado, em ``data/processed/``.
NOME_DICIONARIO = "dicionario_colunas_comuns.csv"


def dicionario_de_dados(
    anos: Sequence[int],
    *,
    apenas_comuns: bool = True,
    linhas: int | None = None,
    incluir_descricao: bool = True,
    paths: Paths | None = None,
) -> pd.DataFrame:
    """Uma linha por variável, descrevendo-a e dizendo se ela serve à série histórica.

    Junta as três fontes que respondem a isso: a presença de cada variável nos
    cabeçalhos de cada ano, o perfil dos valores em cada ano, e a descrição
    oficial da planilha do INEP.

    :param apenas_comuns: restringe às variáveis presentes em todos os anos —
        as únicas comparáveis sem ressalva. Com ``False``, traz todas e marca a
        situação na coluna ``comum``.
    :param linhas: repassado a :func:`perfilar`. ``None`` (padrão) varre os
        arquivos inteiros; é o único modo em que ``dtype_serie`` pode ser levado
        a sério.

    A coluna ``dtype_serie`` é o resultado prático: o dtype que comporta a
    variável em todos os anos. ``tipo_estavel=False`` marca onde os anos
    discordam entre si — é ali que uma concatenação silenciosamente vira
    ``object``.
    """
    anos = sorted(anos)
    presenca = resumo_presenca(anos, paths=paths)
    alvo = list(presenca.loc[presenca["comum"], "coluna"]) if apenas_comuns else None
    if apenas_comuns and not alvo:
        raise ValueError(f"nenhuma coluna é comum a todos os anos {anos}")

    perfis = perfilar_anos(anos, alvo, linhas=linhas, paths=paths)
    tipos = tipos_por_ano(perfis)

    agregado = (
        perfis.groupby("coluna", sort=False)
        .agg(
            natureza=("natureza", "first"),
            dtype_prefixo=("dtype_prefixo", "first"),
            conflito_prefixo=("conflito", "any"),
            # O pior ano manda: uma variável 90% vazia em 2025 não serve à série
            # só porque estava cheia em 2019.
            pct_nulo_max=("pct_nulo", "max"),
            n_distintos=("n_distintos", "max"),
            contagem_exata=("contagem_exata", "all"),
            minimo=("minimo", "min"),
            maximo=("maximo", "max"),
            largura_max=("largura_max", "max"),
            zeros_a_esquerda=("zeros_a_esquerda", "any"),
            exemplos=("exemplos", "last"),
        )
        .reset_index()
    )

    agregado["dtype_serie"] = [promover(tipos.loc[c]) for c in agregado["coluna"]]
    agregado["tipo_estavel"] = [
        len({t for t in tipos.loc[c] if t != VAZIO and pd.notna(t)}) <= 1
        for c in agregado["coluna"]
    ]
    agregado["tipos_por_ano"] = [
        ""
        if estavel
        else ", ".join(f"{ano}:{tipo}" for ano, tipo in tipos.loc[coluna].items() if pd.notna(tipo))
        for coluna, estavel in zip(agregado["coluna"], agregado["tipo_estavel"], strict=True)
    ]

    saida = presenca.merge(agregado, on="coluna", how="right")
    if incluir_descricao:
        saida = saida.merge(descricoes(anos, paths=paths), on="coluna", how="left")
        # Variável sem verbete na planilha fica com descrição vazia, não NaN: o
        # dicionário é para ser lido, e "nan" no meio do texto confunde.
        for campo in ("descricao", "tipo_inep", "tamanho_inep", "categorias"):
            saida[campo] = saida[campo].fillna("")

    ordem = [
        "coluna",
        "descricao",
        "natureza",
        "categorias",
        "dtype_serie",
        "tipo_estavel",
        "tipos_por_ano",
        "dtype_prefixo",
        "conflito_prefixo",
        "tipo_inep",
        "tamanho_inep",
        "comum",
        "n_anos",
        "continua",
        "primeiro_ano",
        "ultimo_ano",
        "anos_presentes",
        "anos_ausentes",
        "pct_nulo_max",
        "n_distintos",
        "contagem_exata",
        "minimo",
        "maximo",
        "largura_max",
        "zeros_a_esquerda",
        "exemplos",
    ]
    return saida[[c for c in ordem if c in saida.columns]]


def dtypes_para_serie_historica(dicionario: pd.DataFrame) -> dict[str, str]:
    """Mapa ``coluna -> dtype`` pronto para o ``dtype=`` de uma leitura.

    Sai do dicionário e entra no pandas sem intermediário: é o produto que o
    documento de preparação existe para gerar.
    """
    return {
        linha.coluna: linha.dtype_serie
        for linha in dicionario.itertuples()
        if linha.dtype_serie != VAZIO
    }


def salvar_dicionario(
    dicionario: pd.DataFrame,
    caminho: Path | str | None = None,
    *,
    paths: Paths | None = None,
) -> Path:
    """Grava o dicionário em CSV (UTF-8, separado por vírgula).

    Vírgula e UTF-8 de propósito, ao contrário dos microdados do INEP: este
    arquivo é nosso, é para ser lido por gente e por planilha, e não há motivo
    para herdar o ``latin-1`` de lá.
    """
    paths = (paths or get_paths()).criar()
    destino = Path(caminho) if caminho else paths.processed / NOME_DICIONARIO
    destino.parent.mkdir(parents=True, exist_ok=True)
    dicionario.to_csv(destino, index=False, encoding="utf-8")
    return destino


def carregar_dicionario(
    caminho: Path | str | None = None,
    *,
    paths: Paths | None = None,
) -> pd.DataFrame:
    """Lê de volta o CSV gravado por :func:`salvar_dicionario`."""
    paths = paths or get_paths()
    origem = Path(caminho) if caminho else paths.processed / NOME_DICIONARIO
    if not origem.exists():
        raise FileNotFoundError(
            f"Dicionário não encontrado em {origem}.\nRode: censo dicionario"
        )
    return pd.read_csv(origem, encoding="utf-8", keep_default_na=False, na_values=[""])
